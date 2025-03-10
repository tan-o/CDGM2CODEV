import openpyxl
import os
import shutil
import CDGM2XML as C2X
import CDGM2CFD as C2CFD
import CDGM2DAT as C2D
import CDGM2DNDT as C2NT


def load_workbook_safely(filename):
    try:
        # 尝试打开文件
        workbook = openpyxl.load_workbook(filename)
        return workbook
    except FileNotFoundError:
        # 文件不存在时抛出错误
        raise FileNotFoundError(f"文件 {filename} 不存在，请检查文件路径是否正确。")
    except openpyxl.utils.exceptions.InvalidFileException:
        # 文件格式无效时抛出错误
        raise ValueError(f"文件 {filename} 格式无效或损坏。")
    except PermissionError:
        # 权限不足时抛出错误
        raise PermissionError(f"权限不足，无法打开文件 {filename}。")
    except Exception as e:
        # 捕获其他未知异常
        raise Exception(f"打开文件 {filename} 时发生错误: {str(e)}")


def list_files_and_select() -> list:

    clfName: str = None
    dndtName: str = None
    # 获取当前文件夹中的所有文件
    files = [f for f in os.listdir('.') if os.path.isfile(f)]

    # 如果没有文件，提示用户
    if not files:
        print("当前文件夹中没有文件。")
        return None

    # 显示文件列表
    print("当前文件夹中的文件列表：")
    for index, file in enumerate(files, start=1):
        print(f"{index}. {file}")
    for item in files:
        if 'CatalogFull' in item:
            clfName = item
        elif 'dndt_constants' in item:
            dndtName = item
        if clfName is not None and dndtName is not None:
            break
    if clfName is None or dndtName is None:
        print("必须存在 CatalogFull 和 dndt_constants 文件才能继续。")
        return None
    # 提示用户选择
    while True:
        try:
            choice = int(input("请输入文件前面的序号选择你要转换的表格（输入0退出）："))
            if choice == 0:
                print("已取消选择。")
                return None
            if 1 <= choice <= len(files):
                selected_file = files[choice - 1]
                break
            else:
                print("无效的编号，请重新输入。")
        except ValueError:
            print("请输入一个有效的数字。")

    return [selected_file, clfName, dndtName]


if __name__ == "__main__":
    os.system('cls' if os.name == 'nt' else 'clear')
    print("""
# #########################################################
# 免责声明:                     
# 本脚本用于把CDGM玻璃库数据转换
# 成CODEV可用格式,数据的填入是按
# 照我自己的理解填的目前用起来并
# 无大碍,但为了保险起见建议在用完
# 后放到ZEMAX中验证            
# 使用说明:
# 该脚本同目录下需要文件“CatalogFull”和“dndt_constants”
# 这两个文件可以去官网Synopsys下载             
# 把成功生成在succeed中的文件
# “CDGM”放到CODEV安装目录下的glass文件夹
# 其余文件放到CODEV安装目录下的macro文件夹中
# 进行覆盖操作前建议备份原文件
# 如果真的不幸丢失源文件，则可以去官网Synopsys下载
# ##########################################################
""")
    isTrue = input("是否同意上述声明[y/n]")
    if isTrue == 'y' or isTrue == 'Y' or isTrue == 'yes':
        filename = list_files_and_select()
        # print(filename)
        if filename is not None:
            folder_name = 'succeed'
            if os.path.exists(folder_name):
                shutil.rmtree(folder_name)
                print(f"已删除已有的文件夹 '{folder_name}' 及其内容")
            os.makedirs(folder_name)
            workbook = load_workbook_safely(filename[0])
            C2X.toXML(filename[0], workbook)
            print("\n转换为XML成功！")
            C2CFD.toCF_dat(filename[0], filename[1], workbook)
            print("写入CatalogFull成功！")
            C2D.writeDAT(workbook)
            print("创建mycdgm成功！")
            C2NT.writeDNDT(filename[2], workbook)
            print("写入dndt_constants成功！")

    input("按任意键退出...")
