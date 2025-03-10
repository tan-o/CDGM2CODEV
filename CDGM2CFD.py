from openpyxl import Workbook
import pandas as pd
import re


def toCF_dat(filename: str, datName: str, workbook: Workbook):
    sheet = workbook['optical glass']
    dfCatalog = pd.read_csv(datName, sep='\t')
    filename = filename if '.xlsx' not in filename else filename.replace(
        '.xlsx', '')
    createTime = ''.join(re.findall(r'\d+', filename))
    # 获取H-K9L的价格
    rowHK9L = 0
    for row_index in range(1, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=1).value == 'H-K9L':
            rowHK9L = row_index

    # 循环开始
    i = 3
    while sheet.cell(i, 1).value != 'over!':
        # 遍历搜索第167列到133的最大与最低波长
        j = 167
        while sheet.cell(i, j).value is None:
            j = j-1
            if j == 96:
                i = i+1
                break
        x = j
        while sheet.cell(i, j).value is not None:
            j = j-1
            if j == 96:
                i = i+1
                break
        new_data = {
            'Catalog': ['MYCDGM'],
            'Glass': [sheet.cell(i, 1).value.replace('-', '').upper() if '-' in sheet.cell(i, 1).value else sheet.cell(i, 1).value.upper()],
            'Index': [sheet.cell(i, 14).value],
            'V-value': [sheet.cell(i, 24).value],
            'NF-NC': [sheet.cell(i, 26).value],
            'DPF': [int(sheet.cell(i, 62).value*10000)],
            'Price': [(sheet.cell(i, 231).value/sheet.cell(rowHK9L, 231).value)*10],
            'Avail': [4],
            'Bubble': ['--'],
            'Stain': ['--'],
            'Low WL': [sheet.cell(2, x).value],
            'High WL': [sheet.cell(2, j+1).value],
            'Price_gob': [(sheet.cell(3, 231).value/sheet.cell(rowHK9L, 231).value)*10],
            'Specific_grav': [sheet.cell(i, 89).value],
            'Manufacture_date': [f"01/{createTime[4:6]}/{createTime[2:4]}"],
            'date_received': [f"01/{createTime[4:6]}/{createTime[2:4]}"],
            'date_index_changed': [f"01/{createTime[4:6]}/{createTime[2:4]}"],
            'Equation_type': ['SCHT' if sheet.cell(i, 28).value is None else 'SSHT'],
            'Replacement': ['--'],
            'Low_exp_coeff': [sheet.cell(i, 78).value/10 if sheet.cell(i, 78).value is not None else '--'],
            'High_exp_coeff': [sheet.cell(i, 79).value/10 if sheet.cell(i, 78).value is not None else '--'],
            'Transform_temp': [sheet.cell(i, 73).value if sheet.cell(i, 78).value is not None else '--'],
            'Knoop': ['--'],
            'Poisson': ['--'],
            'Young': ['--'],
            'Acid_resist': ['--'],
            'Acid_resist_powder': ['--'],
            'Acid_resist_surface': ['--'],
            'Rigidity': ['--'],
            'Water_resist': ['--']
        }
        # 创建新 DataFrame
        new_df = pd.DataFrame(new_data)
        # 使用 pd.concat 将新数据添加到原始 DataFrame
        dfCatalog = pd.concat([dfCatalog, new_df], ignore_index=True)
        # print(i)
        i = i+1
    dfCatalog.to_csv('succeed/CatalogFull.dat', sep='\t', index=False)
