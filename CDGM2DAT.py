from openpyxl import Workbook
import pandas as pd


def writeDAT(workbook: Workbook):
    sheet = workbook['optical glass']
    # 获取H-K9L的价格
    rowHK9L = 0
    for row_index in range(1, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=1).value == 'H-K9L':
            rowHK9L = row_index

    data = {
        'Catalog': ['MYCDGM'],
        'Glass': [sheet.cell(3, 1).value.replace('-', '').upper() if '-' in sheet.cell(3, 1).value else sheet.cell(3, 1).value.upper()],
        'Index': [sheet.cell(3, 14).value],
        'V-value': [sheet.cell(3, 24).value],
        'NF-NC': [sheet.cell(3, 26).value],
        'DPF': [int(sheet.cell(3, 62).value*10000)],
        'Price': [(sheet.cell(3, 231).value/sheet.cell(rowHK9L, 231).value)*10],
        'Avail': [4],
        'Bubble': ['--'],
        'Stain': ['--']
    }
    df = pd.DataFrame(data)
    # 循环开始
    i = 4
    while sheet.cell(i, 1).value != 'over!':
        new_data = {
            'Catalog': ['MYCDGM'],
            'Glass': [sheet.cell(i, 1).value.replace('-', '').upper() if '-' in sheet.cell(i, 1).value else sheet.cell(i, 1).value.upper()],
            'Index': [sheet.cell(i, 14).value],
            'V-value': [sheet.cell(i, 24).value],
            'NF-NC': [sheet.cell(i, 26).value],
            'DPF': [int(sheet.cell(i, 62).value*10000)],
            'Price': [(sheet.cell(i, 231).value/sheet.cell(rowHK9L, 231).value)*10],
            'Avail': [4],
            'Bubble': ['--'],
            'Stain': ['--']
        }
        # 创建新 DataFrame
        new_df = pd.DataFrame(new_data)
        # 使用 pd.concat 将新数据添加到原始 DataFrame
        df = pd.concat([df, new_df], ignore_index=True)
        i = i+1
    df.to_csv('succeed/mycdgm.dat', sep='\t', index=False)
