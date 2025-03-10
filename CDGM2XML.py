import xmltodict
from openpyxl import Workbook
import re


def toXML(filename: str, workbook: Workbook):
    sheet = workbook['optical glass']
    filename = filename if '.xlsx' not in filename else filename.replace(
        '.xlsx', '')
    createTime = ''.join(re.findall(r'\d+', filename))
    glassDict = {
        'Glass': []
    }
    # 获取H-K9L的价格
    rowHK9L = 0
    for row_index in range(1, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=1).value == 'H-K9L':
            rowHK9L = row_index
    # 循环开始
    i = 3
    while sheet.cell(i, 1).value != 'over!':
        glassItem = {
            'GlassName': sheet.cell(i, 1).value.replace('-', '').upper() if '-' in sheet.cell(i, 1).value else sheet.cell(i, 1).value.upper(),
            'NumericName': sheet.cell(i, 2).value
        }

        # 判断equationtype的方法，第28列中的K1是否为空
        if sheet.cell(i, 28).value is None:
            glassItem['EquationType'] = 'Laurent'
        else:
            glassItem['EquationType'] = 'Glass Manufacturer Sellmeier'

        # 遍历搜索第167列到133的最大与最低波长
        j = 167
        while sheet.cell(i, j).value is None:
            j = j-1
            if j == 96:
                i = i+1
                break
        glassItem['LowWavelength'] = str(sheet.cell(2, j).value)
        while sheet.cell(i, j).value is not None:
            j = j-1
            if j == 96:
                i = i+1
                break
        glassItem['HighWavelength'] = str(sheet.cell(2, j+1).value)
        glassItem['DollarSlabPrice'] = str(
            (sheet.cell(i, 231).value/sheet.cell(rowHK9L, 231).value)*10)
        glassItem['DollarStripPrice'] = str(
            (sheet.cell(i, 231).value/sheet.cell(rowHK9L, 231).value)*10)

        # 以录入时间为准，我也不知道在哪里查，可能是生产频次？随便填的
        glassItem['ManufactureDate'] = f"01/{createTime[4:6]}/{createTime[2:4]}"
        glassItem['ReceivedDate'] = f"01/{createTime[4:6]}/{createTime[2:4]}"
        glassItem['IndexChangedDate'] = f"01/{createTime[4:6]}/{createTime[2:4]}"

        glassItem['Availability'] = '3'  # 应该是库存的意思，不知道在哪里查，随便写个

        if sheet.cell(i, 28).value is None:
            glassItem['DispersionCoefficients'] = {'Coefficient': ["{0:.10f}".format(
                sheet.cell(i, x).value).rstrip('0').rstrip('.') for x in range(34, 40)]}
        else:
            glassItem['DispersionCoefficients'] = {'Coefficient': ["{0:.10f}".format(
                sheet.cell(i, x).value).rstrip('0').rstrip('.') for x in range(28, 34)]}

        if sheet.cell(i, 78).value is not None:
            glassItem['LowCTE'] = {
                'LowerTemperatureLimit': '-30',
                'UpperTemperatureLimit': '70',
                'Value': str(sheet.cell(i, 78).value/10),
                'ExpansionCoeffFactor': '10e-6'
            }
        if sheet.cell(i, 79).value is not None:
            glassItem['HighCTE'] = {
                'LowerTemperatureLimit': '100',
                'UpperTemperatureLimit': '300',
                'Value': str(sheet.cell(i, 79).value/10),
                'ExpansionCoeffFactor': '10e-6'
            }
        glassItem['ManufacturersProperties'] = {
            'Property': [{
                'Name': 'Specific_grav',
                'Value': str(sheet.cell(i, 89).value),
                'Category': 'Mechanical'
            }, {
                'Name': 'Transform_temp',
                'Value': str(sheet.cell(i, 73).value),
                'Category': 'Thermal'
            }
            ]
        }

        # print(i)
        j = 167
        while sheet.cell(i, j).value is None:
            j = j-1
            if j == 96:
                i = i+1
                break
        if j >= 132:
            j = 167
            glassItem['TransmissionCurves'] = {
                'Curve': {
                    'Category': 'Visible',
                    'Thickness': '10',
                    'Transmission': [{'Wavelength': str(sheet.cell(2, x).value), 'Value': str(0 if sheet.cell(i, x).value is None else sheet.cell(i, x).value)} for x in range(167, 132, -1)]
                }
            }
        elif j < 132 and j >= 96:
            glassItem['TransmissionCurves'] = {
                'Curve': {
                    'Category': 'Visible',
                    'Thickness': '5',
                    'Transmission': [{'Wavelength': str(sheet.cell(2, x).value), 'Value': str(0 if sheet.cell(i, x).value is None else sheet.cell(i, x).value)} for x in range(131, 96, -1)]
                }
            }

        if sheet.cell(i, 46).value is None:
            glassDict['Glass'].append(glassItem)
        else:
            ####
            glassItem['DnDtData'] = {
                'DnDtForCategory': {
                    'Category': 'Visible',
                    'DnDtWavelengthEntry': [{
                        'DnDtWavelength': {
                            'Wavelength': '435.835',
                            'WavelengthDesignation': 'g',
                            'WavelengthCategory': '5'
                        }, 'DnDtRangeEntry': {
                            'DnDt': str(sheet.cell(i, 223).value),
                            'LowerTemperature': '20',
                            'UpperTemperature': '40'
                        }
                    }, {'DnDtWavelength': {
                        'Wavelength': '479.991',
                        'WavelengthDesignation': "F'",
                        'WavelengthCategory': '4'
                    }, 'DnDtRangeEntry': {
                        'DnDt': str(sheet.cell(i, 213).value),
                        'LowerTemperature': '20',
                        'UpperTemperature': '40'
                    }
                    }, {'DnDtWavelength': {
                        'Wavelength': '546.074',
                        'WavelengthDesignation': 'e',
                        'WavelengthCategory': '3'
                    }, 'DnDtRangeEntry': {
                        'DnDt': str(sheet.cell(i, 203).value),
                        'LowerTemperature': '20',
                        'UpperTemperature': '40'
                    }
                    }, {'DnDtWavelength': {
                        'Wavelength': '643.847',
                        'WavelengthDesignation': "C'",
                        'WavelengthCategory': '2'
                    }, 'DnDtRangeEntry': {
                        'DnDt': str(sheet.cell(i, 183).value),
                        'LowerTemperature': '20',
                        'UpperTemperature': '40'
                    }
                    }, {'DnDtWavelength': {
                        'Wavelength': '1013.98',
                        'WavelengthDesignation': 't',
                        'WavelengthCategory': '1'
                    }, 'DnDtRangeEntry': {
                        'DnDt': str(sheet.cell(i, 173).value),
                        'LowerTemperature': '20',
                        'UpperTemperature': '40'
                    }
                    }],
                    'DnDtConstants': {
                        'Lambda': "{0:.16f}".format(sheet.cell(i, 46).value).rstrip('0').rstrip('.'),
                        'Temperature': '20',
                        'DnDt_D0': "{0:.16f}".format(sheet.cell(i, 41).value).rstrip('0').rstrip('.'),
                        'DnDt_D1': "{0:.16f}".format(sheet.cell(i, 42).value).rstrip('0').rstrip('.'),
                        'DnDt_D2': "{0:.16f}".format(sheet.cell(i, 43).value).rstrip('0').rstrip('.'),
                        'DnDt_E0': "{0:.16f}".format(sheet.cell(i, 44).value).rstrip('0').rstrip('.'),
                        'DnDt_E1': "{0:.16f}".format(sheet.cell(i, 45).value).rstrip('0').rstrip('.')
                    }
                }
            }

            glassDict['Glass'].append(glassItem)

        i = i+1
    glassResult = {
        'Catalog': {
            'Name': 'MYCDGM',
            'Fullname': 'MYCDGM',
            'Version': '3',
            'CatalogProperties': {'DispersionTemperatureIsSameForAllGlasses': 'true'}
        }
    }
    glassResult['Catalog']['Glasses'] = glassDict
    xml_data = xmltodict.unparse(glassResult, pretty=True)
    # 将 XML 数据写入文件
    with open('succeed/MYCDGM.xml', 'w', encoding='utf-8') as file:
        file.write(xml_data)
