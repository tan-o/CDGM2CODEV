from openpyxl import Workbook


def writeDNDT(filename, workbook: Workbook):
    sheet = workbook['optical glass']
    # 文件路径
    file_path = "succeed/dndt_constants.dat"
    try:
        # 打开源文件（只读模式）
        with open(filename, "r", encoding="utf-8") as f_src:
            # 打开目标文件（写入模式）
            with open(file_path, "w", encoding="utf-8") as f_dst:
                # 逐行读取并写入
                for line in f_src:
                    f_dst.write(line)
        # print(f"文件已复制到 {file_path}")
    except FileNotFoundError:
        print(f"错误：源文件 {filename} 不存在")
    except Exception as e:
        print(f"发生错误：{e}")

    # 循环开始
    i = 3
    while sheet.cell(i, 1).value != 'over!':
        if sheet.cell(i, 46).value is None:
            i = i+1
            continue
        new_content = ''
        dndtDataDict = {
            'Glass': sheet.cell(i, 1).value.replace('-', '').upper() if '-' in sheet.cell(i, 1).value else sheet.cell(i, 1).value.upper(),
            'Catalog': 'MYCDGM',
            'Formula': 'lau' if sheet.cell(i, 28).value is None else 'gms',
            'A1/B1': "{:.12E}".format(sheet.cell(i, 28).value) if sheet.cell(i, 28).value is not None else "{:.12E}".format(sheet.cell(i, 34).value),
            'A2/C1': "{:.12E}".format(sheet.cell(i, 29).value) if sheet.cell(i, 28).value is not None else "{:.12E}".format(sheet.cell(i, 35).value),
            'A3/B2': "{:.12E}".format(sheet.cell(i, 30).value) if sheet.cell(i, 28).value is not None else "{:.12E}".format(sheet.cell(i, 36).value),
            'A4/C2': "{:.12E}".format(sheet.cell(i, 31).value) if sheet.cell(i, 28).value is not None else "{:.12E}".format(sheet.cell(i, 37).value),
            'A5/B3': "{:.12E}".format(sheet.cell(i, 32).value) if sheet.cell(i, 28).value is not None else "{:.12E}".format(sheet.cell(i, 38).value),
            'A6/C3': "{:.12E}".format(sheet.cell(i, 33).value) if sheet.cell(i, 28).value is not None else "{:.12E}".format(sheet.cell(i, 39).value),
            'A7/B4': '\t',
            'A8/C4': '\t',
            'A9/': '\t',
            'D0': "{:.12E}".format(sheet.cell(i, 41).value),
            'D1': "{:.12E}".format(sheet.cell(i, 42).value),
            'D2': "{:.12E}".format(sheet.cell(i, 43).value),
            'E0': "{:.12E}".format(sheet.cell(i, 44).value),
            'E1': "{:.12E}".format(sheet.cell(i, 45).value),
            'lambda': "{:.12E}".format(sheet.cell(i, 46).value),
            'MeasuredTemp': '20',
            'Alpha_lowT': '0' if sheet.cell(i, 78).value is None else str(sheet.cell(i, 78).value/10)+'E-6',
            'Tmin_al': '-' if sheet.cell(i, 78).value is None else '-30',
            'Tmax_al': '-' if sheet.cell(i, 78).value is None else '70',
            'Alpha_highT': '0' if sheet.cell(i, 79).value is None else str(sheet.cell(i, 79).value/10)+'E-6',
            'Tmin_ah': '100',
            'Tmax_ah': '300',
            'T0CTEg': '\t',
            'CTEg0': '\t',
            'CTEg1': '\t',
            'CTEg2': '\t',
            'CTEg3': '\t',
            'CTEg4': '\t',
            'CTEg5': '\t',
        }
        for item in dndtDataDict.keys():
            # print(dndtDataDict[item])
            if dndtDataDict[item] == '\t':
                new_content = new_content+dndtDataDict[item]
                continue
            new_content = new_content+dndtDataDict[item]+'\t'
        new_content = new_content+'\n'
        # 以追加模式打开文件并写入
        with open(file_path, "a", encoding="utf-8") as file:
            file.write(new_content)
        # print(i)
        i = i+1
